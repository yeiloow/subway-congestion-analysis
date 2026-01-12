#!/usr/bin/env python3
"""
Utility to check and convert file encoding from euc-kr to utf-8.
Supports CSV and XLSX files.
"""

import os
import sys
from pathlib import Path
from typing import Optional, List


def detect_encoding(file_path: str) -> Optional[str]:
    """
    Detect the encoding of a file using chardet.
    Returns the detected encoding or None if detection fails.
    """
    try:
        import chardet
    except ImportError:
        print("chardet not installed. Installing: pip install chardet")
        os.system("pip install chardet")
        import chardet

    with open(file_path, 'rb') as f:
        raw_data = f.read(100000)  # Read first 100KB for detection
        result = chardet.detect(raw_data)
        return result.get('encoding')


def convert_csv_to_utf8(file_path: str) -> bool:
    """
    Convert CSV file from euc-kr to utf-8 if needed.
    Returns True if conversion was performed, False otherwise.
    """
    detected_encoding = detect_encoding(file_path)

    if not detected_encoding:
        print(f"⚠️  Could not detect encoding for {file_path}")
        return False

    detected_encoding_lower = detected_encoding.lower()

    # Check if file is euc-kr encoded
    if 'euc-kr' in detected_encoding_lower or 'euc_kr' in detected_encoding_lower or 'euckr' in detected_encoding_lower:
        print(f"✓ {file_path} detected as {detected_encoding}, converting to UTF-8...")

        try:
            # Read with euc-kr encoding
            with open(file_path, 'r', encoding='euc-kr') as f:
                content = f.read()

            # Write with UTF-8 encoding
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)

            print(f"✓ Successfully converted {file_path} to UTF-8")
            return True
        except Exception as e:
            print(f"✗ Error converting {file_path}: {e}")
            return False
    else:
        print(f"ℹ {file_path} is already in {detected_encoding}, no conversion needed")
        return False


def convert_xlsx_to_csv(file_path: str) -> bool:
    """
    Convert XLSX file to CSV with UTF-8 encoding.
    Returns True if conversion was performed, False otherwise.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl not installed. Installing: pip install openpyxl")
        os.system("pip install openpyxl")
        from openpyxl import load_workbook

    import csv

    try:
        print(f"✓ Converting XLSX to CSV: {file_path}")

        # Load workbook
        wb = load_workbook(file_path)
        ws = wb.active

        # Generate output CSV file path
        csv_file_path = file_path.replace('.xlsx', '.csv').replace('.XLSX', '.csv')

        # Write to CSV with UTF-8 encoding
        with open(csv_file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        print(f"✓ Successfully converted {file_path} to {csv_file_path}")
        return True
    except Exception as e:
        print(f"✗ Error converting {file_path}: {e}")
        return False


def process_files(directory: str = '.', file_paths: Optional[List[str]] = None) -> None:
    """
    Process CSV and XLSX files in the specified directory or file list.

    Args:
        directory: Directory to search for files (default: current directory)
        file_paths: Specific file paths to process (overrides directory search)
    """
    files_to_process = []

    if file_paths:
        # Process specific files
        files_to_process = file_paths
    else:
        # Find all CSV and XLSX files in directory
        for ext in ['.csv', '.xlsx']:
            files_to_process.extend(Path(directory).glob(f'*{ext}'))
            files_to_process.extend(Path(directory).glob(f'**/*{ext}'))

    if not files_to_process:
        print(f"No CSV or XLSX files found in {directory}")
        return

    print(f"\nProcessing {len(files_to_process)} file(s)...\n")

    converted_count = 0
    for file_path in files_to_process:
        file_path = str(file_path)

        if file_path.lower().endswith('.csv'):
            if convert_csv_to_utf8(file_path):
                converted_count += 1
        elif file_path.lower().endswith('.xlsx'):
            if convert_xlsx_to_csv(file_path):
                converted_count += 1
        print()

    print(f"\n{'='*50}")
    print(f"Processing complete: {converted_count} file(s) converted")
    print(f"{'='*50}")


if __name__ == '__main__':
    if len(sys.argv) > 1:
        # Process specific files passed as arguments
        process_files(file_paths=sys.argv[1:])
    else:
        # Process all CSV and XLSX files in current directory
        process_files()
